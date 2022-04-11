import os.path
import argparse
import cv2
import requests
import json
from tqdm import tqdm
# https://runebook.dev/ko/docs/django/ref/request-response
def getSingle(opt):
    fileName=opt.fileName
    origin_path=opt.origin_path
    roi_path=opt.roi_path
    jsonPath = 'C:/Users/dgdgk/Documents/timeCheck'
    print(origin_path,os.path.join(origin_path,fileName))
    img=cv2.imread(os.path.join(origin_path,fileName))
    rows,cols,_=img.shape
    data={'rows': str(rows), 'cols': str(cols)}
    files = {'media': open(os.path.join(roi_path,fileName), 'rb')}
    print(type(files))

    # response=requests.post("http://203.250.32.251:8000/coocr_upload",files=files)
    # response = requests.post("http://203.250.32.251:8000/api", files=files, cookies=data, params=data)
    response = requests.post("http://203.250.35.243:5920/api", files=files,  params=data)
    j=response.json()
    path=jsonPath
    file_name=fileName.split(".")
    print(fileName,file_name)
    file_name=file_name[0]+".json"
    with open(os.path.join(path,file_name), 'w',encoding='UTF-8') as outfile:
        # json.dump(j, outfile,ensure_ascii=False,indent='\t')
        json.dump(j, outfile, ensure_ascii=False, indent='\t')




def getMulti(opt):
    fileName=opt.fileName
    origin_path=opt.origin_path
    roi_path=opt.roi_path
    fileEx = r'.jpg'
    jsonPath='C:/Users/dgdgk/Documents/timeCheck'
    origin_list = [file for file in os.listdir(origin_path) if file.endswith(fileEx)]
    roi_list = [file for file in os.listdir(roi_path) if file.endswith(fileEx)]
    # assert len(origin_list)==len(roi_list), "len(origin_list) : "+str(len(origin_list))+" len(roi_list) : "+str(len(roi_list))

    for i in tqdm(range(len(roi_list))):
        fileName=roi_list[i]

        img=cv2.imread(os.path.join(origin_path,fileName))
        rows,cols,_=img.shape
        data={'rows': str(rows), 'cols': str(cols)}
        # print(data)
        files = {'media': open(os.path.join(roi_path,fileName), 'rb')}
        # print(type(files))

        # response=requests.post("http://203.250.32.251:8000/coocr_upload",files=files)
        response = requests.post("http://203.250.35.243:5920/api", files=files,  params=data)
        j = response.json()
        path = jsonPath
        file_name = fileName.split(".")
        print('\n',fileName, file_name,'\n')
        file_name = file_name[0] + ".json"
        with open(os.path.join(path, file_name), 'w', encoding='UTF-8') as outfile:
            # json.dump(j, outfile,ensure_ascii=False,indent='\t')
            json.dump(j, outfile, ensure_ascii=False, indent='\t')
        # print(response.json())
        # print(response.content)
        # print(response.text)
        # print(response.content)
        # os.system('pause')
from openpyxl import Workbook
def writeData(ws,row,col,data):
    '''

    :param ws:
    :param col:
    :param data:
    :return:ws
    '''
    for i in range(len(data)):
        ws.cell(row,col,data[i])
        col+=1
    return ws
def makeFormat(ws,row,col,nameArr):
    '''
    출력할 포멧을 만드는 역할임

    :param ws: workSheet
    :param col: 출력을 시작할 위치
    :param nameArr: 출력할 이름들
    :return: ws, col
    '''

    for i in range(len(nameArr)):
        ws.cell(row,col,nameArr[i])
        col+=1
    return ws
def writeExcel(jsonPath):
    fileEx = r'.json'
    json_list = [file for file in os.listdir(jsonPath) if file.endswith(fileEx)]
    path = jsonPath
    wb = Workbook()
    ws = wb.active
    row=7
    col=8
    ws=makeFormat(ws,row,col,["fullText",	"fullText+BrandLeft",	"fullText+BrandRight",	"onlyLine",	"brandLeft+line",	"brandRight+line",
                              "fullText_time",	"fullText+BrandLeft_time",	"fullText+BrandRight_time",	"onlyLine_time",	"brandLeft+line_time",	"brandRight+line_time"])
    row+=1
    for i, data in enumerate(json_list):
        print(os.path.join(path, data))
        lis=[]
        with open(os.path.join(path, data), 'r', encoding='UTF-8') as outfile:
            outfile = json.load(outfile)
            lis.append(data)
            lis.append(outfile['1_result'])
            lis.append(outfile['2_result'])
            lis.append(outfile['3_result'])
            lis.append(outfile['4_result'])
            lis.append(outfile['5_result'])
            lis.append(outfile['6_result'])
            lis.append(outfile['1_time'])
            lis.append(outfile['2_time'])
            lis.append(outfile['3_time'])
            lis.append(outfile['4_time'])
            lis.append(outfile['5_time'])
            lis.append(outfile['6_time'])
            ws=writeData(ws,row,col-1,lis)
            row+=1
            print(os.path.join(path,"FileName" + '.xlsx'))
            wb.save(os.path.join(path,"FileName" + '.xlsx'))
    pass


if __name__ == '__main__':
    # python ./api_test.py --single True --fileName 003.jpg
    # python ./api_test.py --multi True
    parser = argparse.ArgumentParser()
    parser.add_argument('--fileName')
    parser.add_argument('--origin_path', default="../../origin/")
    parser.add_argument('--roi_path', default="../../roi/")
    parser.add_argument('--single', default=False,help="True or False")
    parser.add_argument('--multi', default=False,help="True or False")
    opt=parser.parse_args()
    # print((opt.single == "Treu"))
    assert opt.single == "True" or opt.multi == "True", "single or mulit 둘 중 하나는 true여야합니다"
    if(opt.single=="True"):
        assert opt.fileName != None, "fileName을 입력해주세요"
        getSingle(opt)
    else:
        assert opt.fileName == None, "fileName을 입력하지마세요"
        getMulti(opt)
    # jsonPath = 'C:/Users/dgdgk/Documents/timeCheck'
    # writeExcel(jsonPath)




# import os.path
#
# import cv2
# import requests
# import json
#
#
# if __name__ == '__main__':
#
#     # https://runebook.dev/ko/docs/django/ref/request-response
#     fileName='023.jpg'
#     origin_path="../../cosmetic_demo_image"
#     roi_path="../../test_image"
#     img=cv2.imread(os.path.join(origin_path,fileName))
#     rows,cols,_=img.shape
#     data={'rows': str(rows), 'cols': str(cols)}
#     files = {'media': open(os.path.join(roi_path,fileName), 'rb')}
#     print(type(files))
#
#     # response=requests.post("http://203.250.32.251:8000/coocr_upload",files=files)
#     response=requests.post("http://203.250.32.251:8000/api",files=files,cookies=data)
#     j=response.json()
#     path='C:/Users/dgdgk/Documents'
#     file_name=fileName.split(".")
#     print(fileName,file_name)
#     file_name=file_name[0]+".json"
#     with open(os.path.join(path,file_name), 'w',encoding='UTF-8') as outfile:
#         json.dump(j, outfile,ensure_ascii=False)
#     # print(response.json())
#     # print(response.content)
#     # print(response.text)
#     # print(response.content)